import openpyxl
from django.db.models import TimeField
from django.db.models.functions import Cast
from django.db import IntegrityError
from openpyxl.styles import NamedStyle
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import RegistroRacks, RegistroMaquinas, RegistroContometro
from .forms import RegistroForm, RegistroMQ
from django.contrib import messages

def registro_list(request):
    registros = RegistroRacks.objects.all().order_by('-FECHA')[:27]
    regis = RegistroMaquinas.objects.all().order_by('-FECHA')[:27]
    return render(request, 'mi_app/registro_list.html', {'registros': registros})

def registro_racks(request, RACKS=None):
    rackmq = get_object_or_404(RegistroRacks, RACKS=RACKS) if RACKS else None  # Si se proporciona RCKS, obtener el objeto correspondiente
    
    if request.method == 'POST':
        form = RegistroForm(request.POST, instance=rackmq)     
        if form.is_valid():
            try:
                form.save()
                return redirect('visualizar_registros_racks')  # Redirigir a la lista de registros
            except IntegrityError:
                messages.success(request, "Ya se ha registrado información para este rack hoy.")
    else:
        form = RegistroForm(instance=rackmq)  # Instanciar el formulario con datos existentes o vacío si es nuevo
    return render(request, 'mi_app/registro_racks.html', {'form': form, 'rackmq': rackmq})   

def registro_maquinas(request, MAQUINA_DE_FREON=None):
    mqfreon = get_object_or_404(RegistroMaquinas, MAQUINA_DE_FREON=MAQUINA_DE_FREON) if MAQUINA_DE_FREON else None
    if request.method == 'POST':
        formmq = RegistroMQ(request.POST, instance=mqfreon)
        if formmq.is_valid():
            try:
                formmq.save()
                return redirect('visualizar_registros_maquinas')
            except IntegrityError:
                messages.success(request, "Ya se ha registrado información para esta máquina en este horario.")
    else:
        formmq = RegistroMQ(instance=mqfreon)
    return render(request, 'mi_app/registro_mq.html', {'formmq': formmq, 'mqfreon': mqfreon})

def visualizar_registros_racks(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        registros = RegistroRacks.objects.filter(FECHA__range=[fecha_inicio, fecha_fin]).order_by('-FECHA').order_by('-HORA')[:27]
    else:
        registros = RegistroRacks.objects.all().order_by('FECHA').order_by('-FECHA').order_by('-HORA')[:27]

    if request.GET.get('exportar') == 'excel':
        #crea el archivo excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos exportados"
        #agregar los encabezados
        encabezados = ["FECHA", "HORA", "RACKS", "ESTADO DEL CONDENSADOR", "NIVEL DE REFRIGERANTE", "PRESION ALTA PSIG", "PRESION BAJA COMP 1", "PRESION BAJA COMP 2", "CORRIENTE", "VOLTAJE"]
        ws.append(encabezados)

        date_style = NamedStyle(name="date_style", number_format="dd/mm/yyyy")
        #obtener datos del modelo
        #datos = RegistroRacks.objects.all()
        for dato in registros:
            ws.append([
                dato.FECHA,
                dato.HORA,
                dato.RACKS,
                dato.ESTADO_DEL_CONDENSADOR,
                dato.NIVEL_DE_REFRIGERANTE,
                dato.PRESION_ALTA_PSIG,
                dato.PRESION_DE_BAJA_COMP_1,
                dato.PRESION_DE_BAJA_COMP_2,
                dato.CORRIENTE,
                dato.VOLTAJE
            ])
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.style = date_style

        #Configura la respuesta HTTP para descargar el archivo
        response = HttpResponse(
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename = registroracks.xlsx'
        wb.save(response)
        return response
    return render(request, 'mi_app/visualizar_racks.html', {'registros': registros})

    #registros = RegistroRacks.objects.all().order_by('FECHA')  # Ordenar los registros por fecha descendente

def visualizar_registros_maquinas(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        registros = RegistroMaquinas.objects.filter(FECHA__range=[fecha_inicio, fecha_fin]).order_by('-FECHA').order_by('-HORA')[:28]
    else:
        registros = RegistroMaquinas.objects.all().order_by('FECHA').order_by('-HORA')[:28]

    if request.GET.get('exportar') == 'excel':
        #crea el archivo excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos exportados"
        #agregar los encabezados
        encabezados = ["FECHA", "HORA", "MAQUINA DE FREON", "HORARIO", "TEMPERATURA DE AGUA DE INGRESO", "TEMPERATURA DE AGUA DE TINA", "SALINIDAD DE AGUA EN TINA", "PRESION BAJA PSIG", "CORRIENTE", "VOLTAJE", "CONTOMETRO_1", "CONTOMETRO_2", "MOTIVO DE PARADA", "HORAS DE PARADA"]
        ws.append(encabezados)

        date_style = NamedStyle(name="date_style", number_format="dd/mm/yyyy")
        #obtener datos del modelo
        #datos = RegistroMaquinas.objects.all()
        for dato in registros:
            ws.append([
                dato.FECHA,
                dato.HORA,
                dato.MAQUINA_DE_FREON,
                dato.HORARIO,
                dato.TEMPERATURA_DE_AGUA_DE_INGRESO,
                dato.TEMPERATURA_DE_AGUA_DE_TINA,
                dato.SALINIDAD_DE_AGUA_EN_TINA,
                dato.PRESION_BAJA_PSIG,
                dato.CORRIENTE,
                dato.VOLTAJE,
                dato.CONTOMETRO_1,
                dato.CONTOMETRO_2,
                dato.MOTIVO_DE_PARADA,
                dato.HORAS_DE_PARADA
            ])
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.style = date_style
        
        #Configura la respuesta HTTP para descargar el archivo
        response = HttpResponse(
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename = registromaquinas.xlsx'
        wb.save(response)
        return response
    return render(request, 'mi_app/visualizar_mq.html', {'registros': registros})

def reporte_detallado(request):
    # Obtener registros y aplicar filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        registros = RegistroMaquinas.objects.filter(FECHA__range=[fecha_inicio, fecha_fin])
    else:
        registros = RegistroMaquinas.objects.all()
    
    # Ordenar por fecha, máquina y horario específico
    registros = registros.annotate(
        horario_ordenado=Cast('HORARIO', TimeField())
    ).order_by('FECHA', 'MAQUINA_DE_FREON', 'horario_ordenado')

    # Generar el reporte y calcular las toneladas
    reporte = []
    ultimo_contometro = None
    ultimo_contometro2 = None
    ultima_maquina = None  # Para identificar cambios de máquina
    
    for registro in registros:
        # Calcular la diferencia de toneladas
        if registro.MAQUINA_DE_FREON!='CHILLER':
            if ultimo_contometro is not None and registro.MAQUINA_DE_FREON == ultima_maquina:
                tonelada = registro.CONTOMETRO_1 - ultimo_contometro
            else:
                tonelada = 0  # La primera entrada de una máquina no tiene cálculo
            if ultimo_contometro2 is not None and registro.MAQUINA_DE_FREON == ultima_maquina:
                tonelada2 = registro.CONTOMETRO_2 - ultimo_contometro2
            else:
                tonelada2 = 0    
            
            # Agregar el registro con las toneladas calculadas
            reporte.append({
                'fecha': registro.FECHA,
                'maquina': registro.MAQUINA_DE_FREON,
                'hora': registro.HORARIO,
                'contometro': registro.CONTOMETRO_1,
                'contometro2': registro.CONTOMETRO_2,
                'tonelada': tonelada,
                'tonelada2': tonelada2,
            })

        # Actualizar el último valor del contómetro
        ultimo_contometro = registro.CONTOMETRO_1
        ultimo_contometro2 = registro.CONTOMETRO_2
        ultima_maquina = registro.MAQUINA_DE_FREON
    
    
    # Guardar los datos calculados en RegistroContometros
    for item in reporte:
        if not RegistroContometro.objects.filter(
            FECHA=item['fecha'], 
            MAQUINA=item['maquina'], 
            HORARIO=item['hora']
        ).exists():
            RegistroContometro.objects.create(
                FECHA=item['fecha'],
                MAQUINA=item['maquina'],
                HORARIO=item['hora'],
                CONTOMETRO=item['contometro'],
                CONTOMETRO2=item['contometro2'],
                TONELADA=item['tonelada'],
                TONELADA2=item['tonelada2'],
            )
    
    if request.GET.get('exportar') == 'excel':
        # Crea el archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Detallado"
        
        # Agregar encabezados
        encabezados = ["FECHA", "MAQUINA", "HORARIO", "CONTOMETRO 1", "CONTOMETRO 2", "TONELADA", "TONELADA 2"]
        ws.append(encabezados)

        # Agregar los datos de los registros al archivo Excel
        for item in reporte:
            ws.append([
                item['fecha'],
                item['maquina'],
                item['hora'],
                item['contometro'],
                item['contometro2'],
                item['tonelada'],
                item['tonelada2']
            ])

        # Crear el estilo para las fechas
        date_style = NamedStyle(name="date_style", number_format="dd/mm/yyyy")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.style = date_style

        # Configurar la respuesta HTTP para la descarga del archivo
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_detallado.xlsx"'
        wb.save(response)
        return response
    return render(request, 'mi_app/reporte_detallado.html', {'reporte': reporte})